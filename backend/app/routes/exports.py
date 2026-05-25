import io
from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from app.models.advance import AdvanceRequest
from app.models.expense import ExpenseRequest
from app.models.accounting import LedgerEntry, AccountingAccount

# openpyxl imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# reportlab imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.core.dependencies import get_db, get_current_active_user
from app.models.user import User
from app.services.tenant import list_expenses_for_company, list_ledger_entries_for_company, list_cash_transactions_for_company, list_projects_summary_for_company, list_departments_summary_for_company, list_audit_entries_for_company

router = APIRouter(prefix="/exports", tags=["Exports"])


@router.get("/expenses/excel")
def export_expenses_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    department_id: Optional[str] = Query(None),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "status": status,
        "project_id": project_id,
        "department_id": department_id,
        "limit": 10000,  # Large limit to export all
        "page": 1
    }
    
    expenses = list_expenses_for_company(db, current_user, filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes de Frais"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Design Styles
    font_family = "Segoe UI"
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='double', color='000000')
    )
    
    # Status styling maps (Soft pastel backgrounds with matching dark text)
    status_styles = {
        "pending": {
            "fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="92400E")
        },
        "approved": {
            "fill": PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="03543F")
        },
        "rejected": {
            "fill": PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="9B1C1C")
        },
        "reimbursed": {
            "fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="0F5132")
        }
    }
    
    headers = [
        "Date", 
        "Employé", 
        "Code Comptable",
        "Catégorie", 
        "Description", 
        "Montant HT",
        "Montant TVA",
        "Montant TTC",
        "Devise", 
        "Statut", 
        "Date de soumission"
    ]
    
    ws.append(headers)
    
    # Style header row
    ws.row_dimensions[1].height = 26
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    currency = current_user.company.currency if current_user.company else "FCFA"
    amount_format = f'#,##0" {currency}"'
    
    # Populate rows
    row_idx = 2
    total_valid_ht = 0.0
    total_valid_tax = 0.0
    total_valid_ttc = 0.0
    
    for exp in expenses:
        emp_name = exp.user.name if exp.user else str(exp.user_id)
        exp_date = exp.expense_date.strftime("%Y-%m-%d") if exp.expense_date else ""
        sub_date = exp.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if exp.submitted_at else ""
        
        ttc = float(exp.amount)
        tax = float(exp.tax_amount or 0.0)
        ht = ttc - tax
        code_comptable = exp.category_rel.code if exp.category_rel else ""
        
        st_lower = exp.status.lower()
        if st_lower in ["approved", "paid", "reimbursed"]:
            total_valid_ht += ht
            total_valid_tax += tax
            total_valid_ttc += ttc
            
        row_data = [
            exp_date,
            emp_name,
            code_comptable or "",
            exp.category,
            exp.description or "",
            ht,
            tax,
            ttc,
            exp.currency,
            st_lower,
            sub_date
        ]
        ws.append(row_data)
        
        # Style content row
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            
            # Alignments
            if col_num in [1, 3, 9, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [6, 7, 8]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = amount_format
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Status styling and translation
            if col_num == 10:
                st = cell.value
                if st in status_styles:
                    cell.fill = status_styles[st]["fill"]
                    cell.font = status_styles[st]["font"]
                    translations = {
                        "pending": "En attente ⏳",
                        "approved": "Approuvé ✅",
                        "rejected": "Refusé ❌",
                        "reimbursed": "Payé 💳"
                    }
                    cell.value = translations.get(st, st.capitalize())
        row_idx += 1
        
    # Totals Row
    ws.row_dimensions[row_idx].height = 22
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    # Label cell (Description col 5)
    cell_lbl = ws.cell(row=row_idx, column=5, value="TOTAL (Approuvées & Payées)")
    cell_lbl.font = Font(name=font_family, size=10, bold=True)
    cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
    cell_lbl.fill = total_fill
    cell_lbl.border = double_bottom_border
    
    # HT Sum cell (col 6)
    cell_ht = ws.cell(row=row_idx, column=6, value=total_valid_ht)
    cell_ht.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
    cell_ht.alignment = Alignment(horizontal="right", vertical="center")
    cell_ht.number_format = amount_format
    cell_ht.fill = total_fill
    cell_ht.border = double_bottom_border

    # Tax Sum cell (col 7)
    cell_tax = ws.cell(row=row_idx, column=7, value=total_valid_tax)
    cell_tax.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
    cell_tax.alignment = Alignment(horizontal="right", vertical="center")
    cell_tax.number_format = amount_format
    cell_tax.fill = total_fill
    cell_tax.border = double_bottom_border

    # TTC Sum cell (col 8)
    cell_ttc = ws.cell(row=row_idx, column=8, value=total_valid_ttc)
    cell_ttc.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
    cell_ttc.alignment = Alignment(horizontal="right", vertical="center")
    cell_ttc.number_format = amount_format
    cell_ttc.fill = total_fill
    cell_ttc.border = double_bottom_border
    
    # Empty total fillers
    for col_num in range(1, len(headers) + 1):
        if col_num not in [5, 6, 7, 8]:
            cell = ws.cell(row=row_idx, column=col_num, value="")
            cell.fill = total_fill
            cell.border = double_bottom_border
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = f"000,000 {currency}"  # Estimated width for totals
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Set Auto-filters
    ws.auto_filter.ref = f"A1:K{row_idx-1}"
    
    # Save workbook to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=expenses_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )


@router.get("/expenses/pdf")
def export_expenses_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    department_id: Optional[str] = Query(None),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "status": status,
        "project_id": project_id,
        "department_id": department_id,
        "limit": 10000,
        "page": 1
    }
    
    expenses = list_expenses_for_company(db, current_user, filters)
    
    buffer = io.BytesIO()
    # Landscape Letter format to display columns comfortably
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    
    # ReportLab Styles
    styles = getSampleStyleSheet()
    
    brand_indigo = colors.HexColor("#4F46E5")
    text_dark = colors.HexColor("#1F2937")
    bg_zebra = colors.HexColor("#F9FAFB")
    border_light = colors.HexColor("#E5E7EB")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=brand_indigo,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=12
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#6B7280"),
        leading=13
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=text_dark,
        leading=11
    )
    
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    
    comp_name = current_user.company.name if current_user.company else "FlowBon"
    story.append(Paragraph("État Récapitulatif des Notes de Frais", title_style))
    story.append(Paragraph(f"Société : {comp_name}", subtitle_style))
    
    # Metadata Info Block
    date_range_str = "Toutes les dates"
    if from_date and to_date:
        date_range_str = f"Du {from_date} au {to_date}"
    elif from_date:
        date_range_str = f"À partir du {from_date}"
    elif to_date:
        date_range_str = f"Jusqu'au {to_date}"
        
    status_str = status.capitalize() if status else "Tous les statuts"
    
    meta_html = f"""
    <b>Période :</b> {date_range_str}<br/>
    <b>Filtre Statut :</b> {status_str}<br/>
    <b>Généré par :</b> {current_user.name} ({current_user.email})<br/>
    <b>Date de génération :</b> {date.today().strftime('%d %B %Y')}
    """
    story.append(Paragraph(meta_html, meta_style))
    story.append(Spacer(1, 15))
    
    # Table Header Row
    table_headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("Employé", cell_header_style),
        Paragraph("Catégorie", cell_header_style),
        Paragraph("Description", cell_header_style),
        Paragraph("Montant", cell_header_style),
        Paragraph("Statut", cell_header_style)
    ]
    
    table_data = [table_headers]
    
    translations = {
        "pending": "En attente",
        "approved": "Approuvé",
        "rejected": "Refusé",
        "reimbursed": "Payé"
    }
    
    def make_status_p(status_val):
        text_colors = {
            "pending": "#92400E",
            "approved": "#03543F",
            "rejected": "#9B1C1C",
            "reimbursed": "#0F5132"
        }
        fg = text_colors.get(status_val, "#1F2937")
        lbl = translations.get(status_val, status_val.capitalize())
        return Paragraph(f"<font color='{fg}'><b>{lbl}</b></font>", ParagraphStyle(
            'St', parent=cell_style, alignment=1
        ))
        
    total_amount = 0.0
    currency = current_user.company.currency if current_user.company else "FCFA"
    
    # Populate rows
    for exp in expenses:
        emp_name = exp.user.name if exp.user else str(exp.user_id)
        exp_date = exp.expense_date.strftime("%Y-%m-%d") if exp.expense_date else ""
        
        desc = exp.description or ""
        if len(desc) > 55:
            desc = desc[:52] + "..."
            
        amt = float(exp.amount)
        st_lower = exp.status.lower()
        if st_lower in ["approved", "paid", "reimbursed"]:
            total_amount += amt
        
        code_comptable = exp.category_rel.code if exp.category_rel else ""
        cat_disp = f"{exp.category} ({code_comptable})" if code_comptable else exp.category
        
        row = [
            Paragraph(exp_date, ParagraphStyle('CDate', parent=cell_style, alignment=1)),
            Paragraph(emp_name, cell_style),
            Paragraph(cat_disp, cell_style),
            Paragraph(desc, cell_style),
            Paragraph(f"{amt:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2)),
            make_status_p(exp.status.lower())
        ]
        table_data.append(row)
        
    # Totals Row
    total_row = [
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("<b>TOTAL (Approuvées & Payées)</b>", ParagraphStyle('TotLbl', parent=cell_style, alignment=2)),
        Paragraph(f"<b>{total_amount:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_indigo)),
        Paragraph("", cell_style)
    ]
    table_data.append(total_row)
    
    # Layout col widths (Total 720 points for Landscape Letter margins)
    col_widths = [75, 110, 110, 220, 105, 100]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_indigo),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, border_light),
    ]
    
    # Row backgrounds
    for i in range(1, len(table_data) - 1):
        bg = bg_zebra if i % 2 == 1 else colors.white
        t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
        t_style.append(('BOTTOMPADDING', (0, i), (-1, i), 6))
        t_style.append(('TOPPADDING', (0, i), (-1, i), 6))
        
        # Color cell backgrounds in the Status column
        exp = expenses[i-1]
        st = exp.status.lower()
        status_bg = {
            "pending": colors.HexColor("#FEF3C7"),
            "approved": colors.HexColor("#DEF7EC"),
            "rejected": colors.HexColor("#FDE8E8"),
            "reimbursed": colors.HexColor("#D1E7DD")
        }
        bg_col = status_bg.get(st, colors.HexColor("#F3F4F6"))
        t_style.append(('BACKGROUND', (5, i), (5, i), bg_col))
        
    last_row_idx = len(table_data) - 1
    t_style.extend([
        ('LINEABOVE', (0, last_row_idx), (-1, last_row_idx), 1, brand_indigo),
        ('LINEBELOW', (0, last_row_idx), (-1, last_row_idx), 1.5, brand_indigo),
        ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor("#EEF2F6")),
        ('BOTTOMPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
        ('TOPPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
    ])
    
    t.setStyle(TableStyle(t_style))
    story.append(t)
    story.append(Spacer(1, 35))
    
    # Official Signature Block Table
    sig_data = [
        [
            Paragraph("<b>Signature du Demandeur (Employé)</b>", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9, textColor=colors.HexColor("#4B5563"))),
            Paragraph("<b>Signature pour Approbation (Direction)</b>", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9, textColor=colors.HexColor("#4B5563")))
        ],
        [Spacer(1, 40), Spacer(1, 40)],
        [
            Paragraph("Date : ___________________", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9)),
            Paragraph("Date : ___________________", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9))
        ]
    ]
    sig_table = Table(sig_data, colWidths=[360, 360])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=expenses_report_{date.today().strftime('%Y%m%d')}.pdf"
    }
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers=headers
    )


@router.get("/payroll/excel")
def export_payroll_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
):
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et comptables.")
        
    from sqlalchemy import func
    query = db.query(
        User.name.label("employee_name"),
        User.email.label("employee_email"),
        func.count(ExpenseRequest.id).label("expenses_count"),
        func.sum(ExpenseRequest.amount).label("total_amount")
    ).join(User, ExpenseRequest.user_id == User.id)
    
    if current_user.company_id:
        query = query.filter(ExpenseRequest.company_id == current_user.company_id)
        
    query = query.filter(ExpenseRequest.status == "approved")
    
    if from_date:
        query = query.filter(ExpenseRequest.expense_date >= from_date)
    if to_date:
        query = query.filter(ExpenseRequest.expense_date <= to_date)
        
    results = query.group_by(User.id, User.name, User.email).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remboursements de Paie"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald Green for payroll
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='double', color='000000')
    )
    
    headers = ["Employé", "Email", "Nombre de Bons Approuvés", "Montant à Rembourser"]
    ws.append(headers)
    
    ws.row_dimensions[1].height = 26
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    currency = current_user.company.currency if current_user.company else "FCFA"
    amount_format = f'#,##0" {currency}"'
    
    row_idx = 2
    grand_total = 0.0
    for row in results:
        amt = float(row.total_amount or 0.0)
        grand_total += amt
        ws.append([
            row.employee_name,
            row.employee_email,
            row.expenses_count,
            amt
        ])
        
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            if col_num == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = amount_format
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1
        
    # Totals
    ws.row_dimensions[row_idx].height = 22
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    cell_lbl = ws.cell(row=row_idx, column=2, value="TOTAL GENERAL A REMBOURSER")
    cell_lbl.font = Font(name=font_family, size=10, bold=True)
    cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
    cell_lbl.fill = total_fill
    cell_lbl.border = double_bottom_border
    
    cell_sum = ws.cell(row=row_idx, column=4, value=grand_total)
    cell_sum.font = Font(name=font_family, size=10, bold=True, color="10B981")
    cell_sum.alignment = Alignment(horizontal="right", vertical="center")
    cell_sum.number_format = amount_format
    cell_sum.fill = total_fill
    cell_sum.border = double_bottom_border
    
    for col_num in range(1, len(headers) + 1):
        if col_num not in [2, 4]:
            cell = ws.cell(row=row_idx, column=col_num, value="")
            cell.fill = total_fill
            cell.border = double_bottom_border
            
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=payroll_reimbursements_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )


@router.get("/payroll/pdf")
def export_payroll_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
):
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et comptables.")
        
    from sqlalchemy import func
    query = db.query(
        User.name.label("employee_name"),
        User.email.label("employee_email"),
        func.count(ExpenseRequest.id).label("expenses_count"),
        func.sum(ExpenseRequest.amount).label("total_amount")
    ).join(User, ExpenseRequest.user_id == User.id)
    
    if current_user.company_id:
        query = query.filter(ExpenseRequest.company_id == current_user.company_id)
        
    query = query.filter(ExpenseRequest.status == "approved")
    
    if from_date:
        query = query.filter(ExpenseRequest.expense_date >= from_date)
    if to_date:
        query = query.filter(ExpenseRequest.expense_date <= to_date)
        
    results = query.group_by(User.id, User.name, User.email).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    brand_emerald = colors.HexColor("#10B981")
    text_dark = colors.HexColor("#1F2937")
    bg_zebra = colors.HexColor("#F9FAFB")
    border_light = colors.HexColor("#E5E7EB")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=brand_emerald,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=12
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#6B7280"),
        leading=13
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=text_dark,
        leading=11
    )
    
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    
    comp_name = current_user.company.name if current_user.company else "FlowBon"
    story.append(Paragraph("État des Remboursements de Frais (RH / Paie)", title_style))
    story.append(Paragraph(f"Société : {comp_name}", subtitle_style))
    
    date_range_str = "Toutes les dates"
    if from_date and to_date:
        date_range_str = f"Du {from_date} au {to_date}"
    elif from_date:
        date_range_str = f"À partir du {from_date}"
    elif to_date:
        date_range_str = f"Jusqu'au {to_date}"
        
    meta_html = f"""
    <b>Période :</b> {date_range_str}<br/>
    <b>Type d'export :</b> Consolidation mensuelle pour remboursement de paie<br/>
    <b>Généré par :</b> {current_user.name} ({current_user.email})<br/>
    <b>Date :</b> {date.today().strftime('%d %B %Y')}
    """
    story.append(Paragraph(meta_html, meta_style))
    story.append(Spacer(1, 15))
    
    table_headers = [
        Paragraph("Employé", cell_header_style),
        Paragraph("Email", cell_header_style),
        Paragraph("Nombre de Bons", cell_header_style),
        Paragraph("Montant à Rembourser", cell_header_style)
    ]
    
    table_data = [table_headers]
    currency = current_user.company.currency if current_user.company else "FCFA"
    
    grand_total = 0.0
    for row in results:
        amt = float(row.total_amount or 0.0)
        grand_total += amt
        
        table_data.append([
            Paragraph(row.employee_name, cell_style),
            Paragraph(row.employee_email, cell_style),
            Paragraph(str(row.expenses_count), ParagraphStyle('CDate', parent=cell_style, alignment=1)),
            Paragraph(f"{amt:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2))
        ])
        
    total_row = [
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("<b>TOTAL GENERAL</b>", ParagraphStyle('TotLbl', parent=cell_style, alignment=2)),
        Paragraph(f"<b>{grand_total:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_emerald))
    ]
    table_data.append(total_row)
    
    col_widths = [140, 160, 80, 124]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_emerald),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, border_light),
    ]
    
    for i in range(1, len(table_data) - 1):
        bg = bg_zebra if i % 2 == 1 else colors.white
        t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
        t_style.append(('BOTTOMPADDING', (0, i), (-1, i), 6))
        t_style.append(('TOPPADDING', (0, i), (-1, i), 6))
        
    last_row_idx = len(table_data) - 1
    t_style.extend([
        ('LINEABOVE', (0, last_row_idx), (-1, last_row_idx), 1, brand_emerald),
        ('LINEBELOW', (0, last_row_idx), (-1, last_row_idx), 1.5, brand_emerald),
        ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor("#EEF2F6")),
        ('BOTTOMPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
        ('TOPPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
    ])
    
    t.setStyle(TableStyle(t_style))
    story.append(t)
    story.append(Spacer(1, 35))
    
    sig_data = [
        [
            Paragraph("<b>Signature de l'Administration / RH</b>", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9, textColor=colors.HexColor("#4B5563"))),
            Paragraph("<b>Signature pour Approbation (Direction)</b>", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9, textColor=colors.HexColor("#4B5563")))
        ],
        [Spacer(1, 40), Spacer(1, 40)],
        [
            Paragraph("Date : ___________________", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9)),
            Paragraph("Date : ___________________", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9))
        ]
    ]
    sig_table = Table(sig_data, colWidths=[252, 252])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=payroll_reimbursements_report_{date.today().strftime('%Y%m%d')}.pdf"
    }
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers=headers
    )


@router.get("/advances/pdf")
def export_advances_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
):
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs.")
        
    query = db.query(AdvanceRequest).options(
        joinedload(AdvanceRequest.user),
        joinedload(AdvanceRequest.expenses)
    )
    
    if current_user.company_id:
        query = query.filter(AdvanceRequest.company_id == current_user.company_id)
        
    if status:
        query = query.filter(AdvanceRequest.status == status)
    else:
        # Include disbursed and reconciled by default to keep report highly useful
        query = query.filter(AdvanceRequest.status.in_(["disbursed", "reconciled"]))
        
    if from_date:
        query = query.filter(AdvanceRequest.created_at >= from_date)
    if to_date:
        query = query.filter(AdvanceRequest.created_at <= to_date)
        
    advances = query.order_by(AdvanceRequest.created_at.desc()).all()
    
    buffer = io.BytesIO()
    # Landscape Letter format to accommodate columns cleanly
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    brand_amber = colors.HexColor("#D97706") # Amber theme
    text_dark = colors.HexColor("#1F2937")
    bg_zebra = colors.HexColor("#FDFAF7") # warm warm zebra
    border_light = colors.HexColor("#E5E7EB")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=brand_amber,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=12
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#6B7280"),
        leading=13
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        textColor=text_dark,
        leading=11
    )
    
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        textColor=colors.white,
        alignment=1
    )
    
    comp_name = current_user.company.name if current_user.company else "FlowBon"
    story.append(Paragraph("Bilan et Réconciliation des Avances de Caisse", title_style))
    story.append(Paragraph(f"Société : {comp_name}", subtitle_style))
    
    date_range_str = "Toutes les dates"
    if from_date and to_date:
        date_range_str = f"Du {from_date} au {to_date}"
    elif from_date:
        date_range_str = f"À partir du {from_date}"
    elif to_date:
        date_range_str = f"Jusqu'au {to_date}"
        
    status_str = status.capitalize() if status else "Avances Actives & Clôturées"
    
    meta_html = f"""
    <b>Période :</b> {date_range_str}<br/>
    <b>Filtre Statut :</b> {status_str}<br/>
    <b>Généré par :</b> {current_user.name} ({current_user.email})<br/>
    <b>Date :</b> {date.today().strftime('%d %B %Y')}
    """
    story.append(Paragraph(meta_html, meta_style))
    story.append(Spacer(1, 15))
    
    table_headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("Employé", cell_header_style),
        Paragraph("Motif / Description", cell_header_style),
        Paragraph("Montant Initial", cell_header_style),
        Paragraph("Montant Justifié", cell_header_style),
        Paragraph("Solde Restant", cell_header_style),
        Paragraph("Statut", cell_header_style)
    ]
    
    table_data = [table_headers]
    currency = current_user.company.currency if current_user.company else "FCFA"
    
    total_initial = 0.0
    total_justified = 0.0
    total_balance = 0.0
    
    translations = {
        "draft": "Brouillon",
        "pending": "En attente",
        "approved": "Approuvé",
        "disbursed": "En cours 💰",
        "rejected": "Refusé",
        "reconciled": "Réconcilié ✅"
    }
    
    def make_status_p(val):
        text_colors = {
            "disbursed": "#92400E",
            "reconciled": "#03543F"
        }
        fg = text_colors.get(val, "#1F2937")
        lbl = translations.get(val, val.capitalize())
        return Paragraph(f"<font color='{fg}'><b>{lbl}</b></font>", ParagraphStyle(
            'St', parent=cell_style, alignment=1
        ))
        
    for adv in advances:
        emp_name = adv.user.name if adv.user else str(adv.user_id)
        adv_date = adv.disbursed_at.strftime("%Y-%m-%d") if adv.disbursed_at else adv.created_at.strftime("%Y-%m-%d")
        
        desc = adv.description or ""
        if len(desc) > 55:
            desc = desc[:52] + "..."
            
        initial = float(adv.amount)
        justified = sum(float(exp.amount) for exp in adv.expenses if exp.status.lower() in ["approved", "paid", "reimbursed"])
        balance = initial - justified
        
        total_initial += initial
        total_justified += justified
        total_balance += balance
        
        table_data.append([
            Paragraph(adv_date, ParagraphStyle('CDate', parent=cell_style, alignment=1)),
            Paragraph(emp_name, cell_style),
            Paragraph(desc, cell_style),
            Paragraph(f"{initial:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2)),
            Paragraph(f"{justified:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2)),
            Paragraph(f"{balance:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2)),
            make_status_p(adv.status.lower())
        ])
        
    total_row = [
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("<b>TOTAL</b>", ParagraphStyle('TotLbl', parent=cell_style, alignment=2)),
        Paragraph(f"<b>{total_initial:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_amber)),
        Paragraph(f"<b>{total_justified:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_amber)),
        Paragraph(f"<b>{total_balance:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_amber)),
        Paragraph("", cell_style)
    ]
    table_data.append(total_row)
    
    col_widths = [75, 115, 185, 95, 95, 95, 60]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_amber),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, border_light),
    ]
    
    for i in range(1, len(table_data) - 1):
        bg = bg_zebra if i % 2 == 1 else colors.white
        t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
        t_style.append(('BOTTOMPADDING', (0, i), (-1, i), 6))
        t_style.append(('TOPPADDING', (0, i), (-1, i), 6))
        
    last_row_idx = len(table_data) - 1
    t_style.extend([
        ('LINEABOVE', (0, last_row_idx), (-1, last_row_idx), 1, brand_amber),
        ('LINEBELOW', (0, last_row_idx), (-1, last_row_idx), 1.5, brand_amber),
        ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor("#FDF6EC")),
        ('BOTTOMPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
        ('TOPPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
    ])
    
    t.setStyle(TableStyle(t_style))
    story.append(t)
    story.append(Spacer(1, 35))
    
    sig_data = [
        [
            Paragraph("<b>Signature de la Trésorerie</b>", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9, textColor=colors.HexColor("#4B5563"))),
            Paragraph("<b>Signature pour Approbation (Direction)</b>", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9, textColor=colors.HexColor("#4B5563")))
        ],
        [Spacer(1, 40), Spacer(1, 40)],
        [
            Paragraph("Date : ___________________", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9)),
            Paragraph("Date : ___________________", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9))
        ]
    ]
    sig_table = Table(sig_data, colWidths=[360, 360])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=advances_reconciliation_report_{date.today().strftime('%Y%m%d')}.pdf"
    }
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers=headers
    )


@router.get("/rejections/pdf")
def export_rejections_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
):
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs.")
        
    from app.models.expense import ExpenseRequest
    from app.models.approval_log import ApprovalLog
    
    query = db.query(ExpenseRequest).options(
        joinedload(ExpenseRequest.user),
        joinedload(ExpenseRequest.category_rel),
        joinedload(ExpenseRequest.approval_logs)
    ).filter(ExpenseRequest.status == "rejected")
    
    if current_user.company_id:
        query = query.filter(ExpenseRequest.company_id == current_user.company_id)
        
    if from_date:
        query = query.filter(ExpenseRequest.expense_date >= from_date)
    if to_date:
        query = query.filter(ExpenseRequest.expense_date <= to_date)
        
    expenses = query.order_by(ExpenseRequest.expense_date.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    brand_rose = colors.HexColor("#E11D48") # Dark Rose/Red theme for Audits & Rejections
    text_dark = colors.HexColor("#1F2937")
    bg_zebra = colors.HexColor("#FFF5F5") # soft rose zebra
    border_light = colors.HexColor("#F3F4F6")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=brand_rose,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=12
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#6B7280"),
        leading=13
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        textColor=text_dark,
        leading=11
    )
    
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        textColor=colors.white,
        alignment=1
    )
    
    comp_name = current_user.company.name if current_user.company else "FlowBon"
    story.append(Paragraph("Registre d'Audit — Bons de Frais Rejetés", title_style))
    story.append(Paragraph(f"Société : {comp_name}", subtitle_style))
    
    date_range_str = "Toutes les dates"
    if from_date and to_date:
        date_range_str = f"Du {from_date} au {to_date}"
    elif from_date:
        date_range_str = f"À partir du {from_date}"
    elif to_date:
        date_range_str = f"Jusqu'au {to_date}"
        
    meta_html = f"""
    <b>Période :</b> {date_range_str}<br/>
    <b>Type d'export :</b> Registre des rejets et anti-fraude<br/>
    <b>Généré par :</b> {current_user.name} ({current_user.email})<br/>
    <b>Date :</b> {date.today().strftime('%d %B %Y')}
    """
    story.append(Paragraph(meta_html, meta_style))
    story.append(Spacer(1, 15))
    
    table_headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("Employé", cell_header_style),
        Paragraph("Catégorie", cell_header_style),
        Paragraph("Description / Motif", cell_header_style),
        Paragraph("Montant TTC", cell_header_style),
        Paragraph("Motif du Rejet (Commentaire)", cell_header_style)
    ]
    
    table_data = [table_headers]
    currency = current_user.company.currency if current_user.company else "FCFA"
    
    total_amount = 0.0
    for exp in expenses:
        emp_name = exp.user.name if exp.user else str(exp.user_id)
        exp_date = exp.expense_date.strftime("%Y-%m-%d")
        cat_name = f"{exp.category_rel.name} ({exp.category_rel.code})" if exp.category_rel else (exp.category or "—")
        
        desc = exp.description or ""
        if len(desc) > 40:
            desc = desc[:37] + "..."
            
        amt = float(exp.amount)
        total_amount += amt
        
        # Get rejection comment
        rejection_comment = "Aucun commentaire fourni"
        reject_logs = [log for log in exp.approval_logs if log.action.lower() == "reject"]
        if reject_logs:
            reject_logs.sort(key=lambda x: x.created_at, reverse=True)
            rejection_comment = reject_logs[0].comment or "Aucun commentaire fourni"
            
        if len(rejection_comment) > 60:
            rejection_comment = rejection_comment[:57] + "..."
            
        table_data.append([
            Paragraph(exp_date, ParagraphStyle('CDate', parent=cell_style, alignment=1)),
            Paragraph(emp_name, cell_style),
            Paragraph(cat_name, cell_style),
            Paragraph(desc, cell_style),
            Paragraph(f"{amt:,.0f} {currency}".replace(",", " "), ParagraphStyle('CAmt', parent=cell_style, alignment=2)),
            Paragraph(f"<font color='#991B1B'><b>{rejection_comment}</b></font>", cell_style)
        ])
        
    total_row = [
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("<b>TOTAL REJETE</b>", ParagraphStyle('TotLbl', parent=cell_style, alignment=2)),
        Paragraph(f"<b>{total_amount:,.0f} {currency}</b>".replace(",", " "), ParagraphStyle('TotAmt', parent=cell_style, alignment=2, textColor=brand_rose)),
        Paragraph("", cell_style)
    ]
    table_data.append(total_row)
    
    col_widths = [75, 115, 115, 120, 95, 200]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_rose),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, border_light),
    ]
    
    for i in range(1, len(table_data) - 1):
        bg = bg_zebra if i % 2 == 1 else colors.white
        t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
        t_style.append(('BOTTOMPADDING', (0, i), (-1, i), 6))
        t_style.append(('TOPPADDING', (0, i), (-1, i), 6))
        
    last_row_idx = len(table_data) - 1
    t_style.extend([
        ('LINEABOVE', (0, last_row_idx), (-1, last_row_idx), 1, brand_rose),
        ('LINEBELOW', (0, last_row_idx), (-1, last_row_idx), 1.5, brand_rose),
        ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor("#FFF1F2")),
        ('BOTTOMPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
        ('TOPPADDING', (0, last_row_idx), (-1, last_row_idx), 8),
    ])
    
    t.setStyle(TableStyle(t_style))
    story.append(t)
    story.append(Spacer(1, 35))
    
    sig_data = [
        [
            Paragraph("<b>Signature de l'Auditeur Interne</b>", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9, textColor=colors.HexColor("#4B5563"))),
            Paragraph("<b>Signature de la Direction Générale</b>", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9, textColor=colors.HexColor("#4B5563")))
        ],
        [Spacer(1, 40), Spacer(1, 40)],
        [
            Paragraph("Date : ___________________", ParagraphStyle('SigL', parent=styles['Normal'], alignment=0, fontSize=9)),
            Paragraph("Date : ___________________", ParagraphStyle('SigR', parent=styles['Normal'], alignment=2, fontSize=9))
        ]
    ]
    sig_table = Table(sig_data, colWidths=[360, 360])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=rejected_expenses_report_{date.today().strftime('%Y%m%d')}.pdf"
    }
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers=headers
    )


@router.get("/attachments/zip")
def export_attachments_zip(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
):
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs.")
        
    from app.models.attachment import Attachment
    from app.models.expense import ExpenseRequest
    from app.config import settings
    from pathlib import Path
    import zipfile
    
    query = db.query(Attachment).join(ExpenseRequest, Attachment.expense_request_id == ExpenseRequest.id)
    
    if current_user.company_id:
        query = query.filter(Attachment.company_id == current_user.company_id)
        
    if from_date:
        query = query.filter(ExpenseRequest.expense_date >= from_date)
    if to_date:
        query = query.filter(ExpenseRequest.expense_date <= to_date)
        
    attachments = query.all()
    
    zip_buffer = io.BytesIO()
    upload_dir = Path(settings.uploads_dir)
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        if not attachments:
            zip_file.writestr("readme.txt", "Aucun justificatif trouve pour la periode selectionnee.")
        else:
            for att in attachments:
                saved_file_name = att.file_url.split("/")[-1]
                physical_path = upload_dir / saved_file_name
                if physical_path.exists():
                    emp_name = att.expense.user.name.replace(" ", "_") if att.expense and att.expense.user else "employe"
                    exp_date = att.expense.expense_date.strftime("%Y%m%d") if att.expense and att.expense.expense_date else "date"
                    zip_filename = f"{emp_name}_{exp_date}_{att.file_name}"
                    zip_file.write(physical_path, zip_filename)
                
    zip_buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=flowbon_attachments_{date.today().strftime('%Y%m%d')}.zip"
    }
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers=headers
    )



@router.get("/ledger/excel")
def export_ledger_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    reference_type: Optional[str] = Query(None),
    account_id: Optional[str] = Query(None),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "reference_type": reference_type,
        "account_id": account_id,
        "limit": 10000,
        "page": 1
    }
    
    entries = list_ledger_entries_for_company(db, current_user, filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Écritures Comptables"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Design Styles
    font_family = "Segoe UI"
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='double', color='000000')
    )
    
    headers = [
        "Référence",
        "Date",
        "Compte",
        "Code",
        "Description",
        "Débit",
        "Crédit",
        "Type",
    ]
    
    ws.append(headers)
    
    # Style header row
    ws.row_dimensions[1].height = 26
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    currency = current_user.company.currency if current_user.company else "FCFA"
    amount_format = f'#,##0" {currency}"'
    
    # Populate rows
    row_idx = 2
    total_debit = 0.0
    total_credit = 0.0
    
    for entry in entries:
        entry_date = entry.transaction_date.strftime("%Y-%m-%d") if entry.transaction_date else ""
        account_name = entry.account.name if entry.account else ""
        account_code = entry.account.code if entry.account else ""
        
        debit = float(entry.debit or 0.0)
        credit = float(entry.credit or 0.0)
        total_debit += debit
        total_credit += credit
        
        row_data = [
            entry.reference_number or "",
            entry_date,
            account_name,
            account_code,
            entry.description or "",
            debit,
            credit,
            entry.reference_type or "",
        ]
        ws.append(row_data)
        
        # Style content row
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            
            # Alignments
            if col_num in [1, 2, 4, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = amount_format
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        row_idx += 1
    
    # Totals Row
    ws.row_dimensions[row_idx].height = 22
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    # Label cell (Description col 5)
    cell_lbl = ws.cell(row=row_idx, column=5, value="TOTAUX")
    cell_lbl.font = Font(name=font_family, size=10, bold=True)
    cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
    cell_lbl.fill = total_fill
    cell_lbl.border = double_bottom_border
    
    # Debit Sum cell (col 6)
    cell_debit = ws.cell(row=row_idx, column=6, value=total_debit)
    cell_debit.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
    cell_debit.alignment = Alignment(horizontal="right", vertical="center")
    cell_debit.number_format = amount_format
    cell_debit.fill = total_fill
    cell_debit.border = double_bottom_border
    
    # Credit Sum cell (col 7)
    cell_credit = ws.cell(row=row_idx, column=7, value=total_credit)
    cell_credit.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
    cell_credit.alignment = Alignment(horizontal="right", vertical="center")
    cell_credit.number_format = amount_format
    cell_credit.fill = total_fill
    cell_credit.border = double_bottom_border
    
    # Empty total fillers
    for col_num in range(1, len(headers) + 1):
        if col_num not in [5, 6, 7]:
            cell = ws.cell(row=row_idx, column=col_num, value="")
            cell.fill = total_fill
            cell.border = double_bottom_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = f"000,000 {currency}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    # Set Auto-filters
    ws.auto_filter.ref = f"A1:H{row_idx-1}"
    
    # Save workbook to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": f"attachment; filename=ledger_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp
    )


@router.get("/ledger/pdf")
def export_ledger_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    reference_type: Optional[str] = Query(None),
    account_id: Optional[str] = Query(None),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "reference_type": reference_type,
        "account_id": account_id,
        "limit": 10000,
        "page": 1
    }
    
    entries = list_ledger_entries_for_company(db, current_user, filters)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.4*72, bottomMargin=0.4*72)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="CustomTitle", parent=styles['Normal'], fontSize=14, textColor=colors.HexColor("#4F46E5"), spaceAfter=12, alignment=1)
    
    story = []
    story.append(Paragraph("Écritures Comptables", title_style))
    
    # Build table data
    table_data = [[
        "Référence",
        "Date",
        "Compte",
        "Code",
        "Description",
        "Débit",
        "Crédit",
        "Type",
    ]]
    
    total_debit = 0.0
    total_credit = 0.0
    
    for entry in entries:
        entry_date = entry.transaction_date.strftime("%Y-%m-%d") if entry.transaction_date else ""
        account_name = entry.account.name if entry.account else ""
        account_code = entry.account.code if entry.account else ""
        debit = float(entry.debit or 0.0)
        credit = float(entry.credit or 0.0)
        total_debit += debit
        total_credit += credit
        
        table_data.append([
            entry.reference_number or "",
            entry_date,
            account_name,
            account_code,
            entry.description or "",
            f"{debit:,.2f}",
            f"{credit:,.2f}",
            entry.reference_type or "",
        ])
    
    # Add totals row
    table_data.append([
        "", "", "", "", "TOTAUX",
        f"{total_debit:,.2f}",
        f"{total_credit:,.2f}",
        ""
    ])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*72, 1*72, 1.5*72, 0.8*72, 1.5*72, 1.2*72, 1.2*72, 1.2*72])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E5E7EB")),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=ledger_export_{date.today().strftime('%Y%m%d')}.pdf"
    }
    
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


# ======================== TREASURY / TRÉSORERIE EXPORTS ========================

@router.get("/treasury/excel")
def export_treasury_excel(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    cash_register_id: Optional[str] = Query(None),
    transaction_type: Optional[str] = Query(None),  # ENTRY or EXIT
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export treasury transactions (cash register movements) as Excel.
    
    Filters:
    - from_date, to_date: Date range
    - cash_register_id: Filter by specific cash register
    - transaction_type: ENTRY or EXIT
    """
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "cash_register_id": cash_register_id,
        "transaction_type": transaction_type,
    }
    
    # Retrieve cash transactions
    transactions = list_cash_transactions_for_company(db, current_user, filters)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trésorerie"
    
    # Define header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column headers
    headers = [
        "Référence",
        "Date",
        "Caisse",
        "Type",
        "Montant",
        "Utilisateur",
        "Description",
        "Référence Bon",
    ]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    column_widths = [1.5, 1.5, 2, 1, 1.5, 1.5, 2.5, 1.5]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width * 72 / 7
    
    # Add data rows
    running_balance = 0
    for trans in transactions:
        # Calculate balance
        if trans.type == "ENTRY":
            running_balance += float(trans.amount)
        else:
            running_balance -= float(trans.amount)
        
        row = [
            trans.reference_number or f"TRX-{trans.id}",
            trans.created_at.strftime("%Y-%m-%d %H:%M") if trans.created_at else "",
            trans.cash_register.name if trans.cash_register else "",
            trans.type,
            float(trans.amount),
            trans.creator.name if trans.creator else "",
            trans.description or "",
            str(trans.reference_id) if trans.reference_id else "",
        ]
        
        ws.append(row)
        
        # Style data row
        for cell in ws[ws.max_row]:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    
    # Return as streaming response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=treasury_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/treasury/pdf")
def export_treasury_pdf(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    cash_register_id: Optional[str] = Query(None),
    transaction_type: Optional[str] = Query(None),  # ENTRY or EXIT
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export treasury transactions (cash register movements) as PDF.
    
    Filters:
    - from_date, to_date: Date range
    - cash_register_id: Filter by specific cash register
    - transaction_type: ENTRY or EXIT
    """
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "cash_register_id": cash_register_id,
        "transaction_type": transaction_type,
    }
    
    # Retrieve cash transactions
    transactions = list_cash_transactions_for_company(db, current_user, filters)
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5 * 72, bottomMargin=0.5 * 72)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=12,
        alignment=1  # center
    )
    story.append(Paragraph("Export Trésorerie", title_style))
    story.append(Spacer(1, 0.3 * 72))
    
    # Prepare table data
    table_data = [["Référence", "Date", "Caisse", "Type", "Montant", "Utilisateur", "Description", "Réf. Bon"]]
    
    for trans in transactions:
        row = [
            trans.reference_number or f"TRX-{trans.id}",
            trans.created_at.strftime("%Y-%m-%d %H:%M") if trans.created_at else "",
            trans.cash_register.name if trans.cash_register else "",
            trans.type,
            f"{float(trans.amount):,.2f}",
            trans.creator.name if trans.creator else "",
            trans.description or "",
            str(trans.reference_id)[:8] if trans.reference_id else "",
        ]
        table_data.append(row)
    
    # Create table
    table = Table(table_data, colWidths=[1.2*72, 1.2*72, 1.5*72, 0.8*72, 1.2*72, 1.2*72, 1.8*72, 0.9*72])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E5E7EB")),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=treasury_export_{date.today().strftime('%Y%m%d')}.pdf"
    }
    
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


# ======================== AGGREGATION EXPORTS (Projects / Departments) ========================

@router.get("/projects/excel")
def export_projects_excel(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export project summary as Excel.
    
    Columns: Project Name, Project Code, Status, Expense Count, Total Amount
    
    Filters:
    - from_date, to_date: Date range for expenses
    - status: Filter by expense status (pending, approved, paid, etc.)
    - include_inactive: Include inactive projects (default: False)
    """
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "status": status,
        "include_inactive": include_inactive,
    }
    
    # Retrieve project summaries
    projects = list_projects_summary_for_company(db, current_user, filters)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projets"
    
    # Define header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column headers
    headers = [
        "Projet",
        "Code",
        "Statut",
        "Nombre de bons",
        "Montant total",
    ]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    column_widths = [2.5, 1.5, 1.5, 1.5, 1.5]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width * 72 / 7
    
    # Add data rows
    for proj in projects:
        row = [
            proj["project_name"],
            proj["project_code"] or "",
            proj["status"],
            proj["expense_count"],
            proj["total_amount"],
        ]
        
        ws.append(row)
        
        # Style data row
        for cell in ws[ws.max_row]:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    
    # Return as streaming response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=projects_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/projects/pdf")
def export_projects_pdf(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "status": status,
        "include_inactive": include_inactive,
    }
    projects = list_projects_summary_for_company(db, current_user, filters)
    currency = current_user.company.currency if current_user.company else "FCFA"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*72, bottomMargin=0.5*72)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Analyse par Projets", ParagraphStyle("T", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#4F46E5"), spaceAfter=12, alignment=1)))
    story.append(Spacer(1, 0.2*72))

    table_data = [["Projet", "Code", "Statut", "Nombre de bons", f"Montant total ({currency})"]]
    for proj in projects:
        table_data.append([
            proj["project_name"],
            proj["project_code"] or "",
            proj["status"],
            str(proj["expense_count"]),
            f"{float(proj['total_amount']):,.0f}".replace(",", " "),
        ])

    table = Table(table_data, colWidths=[2.5*72, 1.2*72, 1.2*72, 1.5*72, 2*72])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=projects_export_{date.today().strftime('%Y%m%d')}.pdf"
    })


@router.get("/departments/excel")
def export_departments_excel(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export department summary as Excel.
    
    Columns: Department Name, Expense Count, Total Amount
    
    Filters:
    - from_date, to_date: Date range for expenses
    - include_inactive: Include inactive departments (default: False)
    """
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "include_inactive": include_inactive,
    }
    
    # Retrieve department summaries
    departments = list_departments_summary_for_company(db, current_user, filters)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Départements"
    
    # Define header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column headers
    headers = [
        "Département",
        "Nombre de dépenses",
        "Montant total",
    ]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    column_widths = [2.5, 1.8, 1.5]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width * 72 / 7
    
    # Add data rows
    for dept in departments:
        row = [
            dept["department_name"],
            dept["expense_count"],
            dept["total_amount"],
        ]
        
        ws.append(row)
        
        # Style data row
        for cell in ws[ws.max_row]:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    
    # Return as streaming response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=departments_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/departments/pdf")
def export_departments_pdf(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "include_inactive": include_inactive,
    }
    departments = list_departments_summary_for_company(db, current_user, filters)
    currency = current_user.company.currency if current_user.company else "FCFA"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*72, bottomMargin=0.5*72)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Analyse par Départements", ParagraphStyle("T", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#4F46E5"), spaceAfter=12, alignment=1)))
    story.append(Spacer(1, 0.2*72))

    table_data = [["Département", "Nombre de bons", f"Montant total ({currency})"]]
    for dept in departments:
        table_data.append([
            dept["department_name"],
            str(dept["expense_count"]),
            f"{float(dept['total_amount']):,.0f}".replace(",", " "),
        ])

    table = Table(table_data, colWidths=[3*72, 2*72, 2.5*72])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=departments_export_{date.today().strftime('%Y%m%d')}.pdf"
    })


# ======================== AUDIT / WORKFLOW EXPORTS ========================

@router.get("/audit/excel")
def export_audit_excel(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export audit trail for expense approvals as Excel.
    
    Columns: Reference, Created By/Date, Manager Approval/Date, Accounting Approval/Date, Payment By/Date, Final Status
    
    Filters:
    - from_date, to_date: Date range for expense creation
    - status: Filter by final expense status (pending, approved, paid, rejected, etc.)
    """
    filters = {
        "from_date": from_date,
        "to_date": to_date,
        "status": status,
    }
    
    # Retrieve audit entries
    audit_entries = list_audit_entries_for_company(db, current_user, filters)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit"
    
    # Define header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column headers
    headers = [
        "Référence Bon",
        "Créé par",
        "Date création",
        "Approuvé Manager",
        "Date approbation Manager",
        "Approuvé Comptable",
        "Date approbation Comptable",
        "Payé par",
        "Date paiement",
        "Statut final",
    ]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    column_widths = [1.5, 1.8, 1.8, 1.8, 2.0, 1.8, 2.0, 1.8, 1.8, 1.2]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width * 72 / 7
    
    # Add data rows
    for entry in audit_entries:
        row = [
            entry["reference_number"],
            entry["created_by"],
            entry["created_at"],
            entry["approved_manager_by"],
            entry["approved_manager_at"],
            entry["approved_accounting_by"],
            entry["approved_accounting_at"],
            entry["paid_by"],
            entry["paid_at"],
            entry["status_final"],
        ]
        
        ws.append(row)
        
        # Style data row
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = left_align
    
    # Return as streaming response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=audit_export_{date.today().strftime('%Y%m%d')}.xlsx"
    }
    
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/audit/pdf")
def export_audit_pdf(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export audit trail as PDF."""
    filters = {"from_date": from_date, "to_date": to_date, "status": status}
    audit_entries = list_audit_entries_for_company(db, current_user, filters)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.4*72, bottomMargin=0.4*72)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(
        "Traçabilité & Audit des Bons",
        ParagraphStyle("T", parent=styles["Heading1"], fontSize=14, textColor=colors.HexColor("#4F46E5"), spaceAfter=12, alignment=1)
    ))
    story.append(Spacer(1, 0.15*72))

    headers_row = [
        "Référence", "Créé par", "Date création",
        "Approbation Mgr", "Date Mgr",
        "Approbation Cpt", "Date Cpt",
        "Payé par", "Date paiement", "Statut"
    ]
    table_data = [headers_row]
    for entry in audit_entries:
        table_data.append([
            entry["reference_number"],
            entry["created_by"],
            entry["created_at"] or "",
            entry["approved_manager_by"] or "",
            entry["approved_manager_at"] or "",
            entry["approved_accounting_by"] or "",
            entry["approved_accounting_at"] or "",
            entry["paid_by"] or "",
            entry["paid_at"] or "",
            entry["status_final"],
        ])

    col_widths = [1.0*72, 1.1*72, 0.9*72, 1.1*72, 0.85*72, 1.1*72, 0.85*72, 1.0*72, 0.85*72, 0.85*72]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=audit_export_{date.today().strftime('%Y%m%d')}.pdf"
    })


@router.get("/fec")
def export_fec(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    fiscal_year_id: Optional[str] = Query(None),
):
    import csv
    from io import StringIO
    from fastapi import HTTPException
    
    if current_user.role not in ["admin", "super_admin", "accountant"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et comptables.")

    query = db.query(LedgerEntry).options(joinedload(LedgerEntry.account))
    query = query.filter(LedgerEntry.company_id == current_user.company_id)

    if fiscal_year_id:
        query = query.filter(LedgerEntry.fiscal_year_id == fiscal_year_id)
        
    entries = query.order_by(LedgerEntry.transaction_date).all()

    output = StringIO()
    writer = csv.writer(output, delimiter='|', lineterminator='\r\n')
    
    headers = [
        "JournalCode", "JournalLib", "EcritureNum", "EcritureDate", 
        "CompteNum", "CompteLib", "CompAuxNum", "CompAuxLib", 
        "PieceRef", "PieceDate", "EcritureLib", "Debit", "Credit", 
        "EcritureLet", "DateLet", "ValidDate", "Montantdevise", "Idevise"
    ]
    writer.writerow(headers)

    for entry in entries:
        dt_str = entry.transaction_date.strftime("%Y%m%d") if entry.transaction_date else ""
        created_str = entry.created_at.strftime("%Y%m%d") if entry.created_at else ""
        
        # Le FEC attend une virgule pour les décimales
        debit_str = f"{float(entry.debit):.2f}".replace('.', ',') if entry.debit else "0,00"
        credit_str = f"{float(entry.credit):.2f}".replace('.', ',') if entry.credit else "0,00"

        row = [
            entry.journal_type or "OD",         # JournalCode
            entry.journal_type or "OD",         # JournalLib
            entry.reference_number or "",       # EcritureNum
            dt_str,                             # EcritureDate
            entry.account.code if entry.account else "", # CompteNum
            entry.account.name if entry.account else "", # CompteLib
            "",                                 # CompAuxNum
            "",                                 # CompAuxLib
            entry.piece_number or "",           # PieceRef
            dt_str,                             # PieceDate
            entry.description or "",            # EcritureLib
            debit_str,                          # Debit
            credit_str,                         # Credit
            "",                                 # EcritureLet
            "",                                 # DateLet
            created_str,                        # ValidDate
            "",                                 # Montantdevise
            ""                                  # Idevise
        ]
        writer.writerow(row)

    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": f"attachment; filename=FEC_{date.today().strftime('%Y%m%d')}.txt"
    }
    
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="text/plain", 
        headers=headers_resp
    )
